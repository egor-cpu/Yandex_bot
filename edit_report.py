import openpyxl
from main import dp, bot, y, greet_kb
import yadisk
from aiogram import Bot, Dispatcher, types
from dotenv import load_dotenv
from aiogram import types
from aiogram.filters import Command
import os
from datetime import datetime
from aiogram.types import ReplyKeyboardRemove, ReplyKeyboardMarkup, KeyboardButton, InlineKeyboardMarkup, InlineKeyboardButton
from aiogram.fsm.state import State, StatesGroup
from aiogram.fsm.context import FSMContext
class editStates(StatesGroup):
    waiting_for_chose_document = State()
    waiting_for_worktype = State()
    waiting_for_chose_row = State()
    waiting_for_element_row = State()
    waiting_for_edit_row = State()
    waiting_for_chose_type_date = State()
    waiting_for_chose_date = State()
    waiting_for_chose_category = State()
    waiting_for_chose_expense_item = State()
    waiting_for_chose_name = State()
    waiting_for_type = State()
    waiting_for_chose_count = State()
    waiting_for_chose_price = State()
    waiting_for_document_numbers = State()
    waiting_for_document_types = State()
    waiting_for_document = State()
    waiting_for_comments = State()
    waiting_for_accept_row = State()
    waiting_for_chose = State()


def convert_date_format(date_text):
    try:
        # Преобразуем текст в объект datetime
        date_obj = datetime.strptime(date_text, "%d.%m.%Y")

        # Форматируем дату в новый формат
        return date_obj.strftime("%d/%m/%Y")
    except ValueError:
        return "Неверный формат даты"

def update_files_from_yandex_disk(local_dir, yandex_disk_dir):
    if not os.path.exists(local_dir):
        os.makedirs(local_dir)

    # Получение списка файлов на Яндекс.Диске
    for item in y.listdir(yandex_disk_dir):
        if item['type'] == 'file':
            file_name = item['name']
            remote_path = f"{yandex_disk_dir}/{file_name}"
            local_path = os.path.join(local_dir, file_name)
            remote_modified_time = item['modified'].timestamp()
            # Проверка необходимости загрузки/обновления файла
            if not os.path.exists(local_path) or remote_modified_time > os.path.getmtime(local_path):
                y.download(remote_path, local_path)
            else:
                print(f"Файл {file_name} уже актуален.")

def get_allrow(line):
    wb = openpyxl.load_workbook("Отчёты/" + filenamesave)
    sheet = wb.active
    alternating_values = ""
    for col in range(1, 11):
        # Берем значение из 1 строки и n строки для текущего столбца
        cell_1 = sheet.cell(row=1, column=col).value
        cell_n = sheet.cell(row=line, column=col).value
        alternating_values = alternating_values + str(cell_1)+": "+ str(cell_n)+"\n"
    return alternating_values

def get_unique_names_from_excel(file_path, sheet_name, column_letter):
    if column_letter==1:
        # Загружаем рабочую книгу и выбираем лист
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook[sheet_name]

        # Создаем множество для хранения уникальных названий
        unique_names = set()

        # Проходим по всем строкам в указанном столбце
        for row in sheet.iter_rows(min_col=column_letter, max_col=column_letter):
            for cell in row:
                if cell.value:  # Проверяем, что ячейка не пустая
                    if str(cell.value) != 'Категория' and str(cell.value) != 'Статья расходов' and str(cell.value) != 'Бюджет':
                        unique_names.add(cell.value)

        return unique_names
    else:
        wb = openpyxl.load_workbook("Отчёты/" + filenamesave)
        sheet1 = wb.active
        cell = sheet1['B' + str(numberline)].value

        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook[sheet_name]
        unique_names = set()
        for row in sheet.iter_rows(min_col=column_letter, max_col=column_letter):
            for cell_in_row in row:
                # Проверка на непустую ячейку и исключение определенных значений
                if cell_in_row.value and str(cell_in_row.value) not in {'Категория', 'Статья расходов', 'Бюджет'}:
                    # Проверка значения в ячейке столбца (column_letter - 1)
                    adjacent_cell = sheet.cell(row=cell_in_row.row, column=column_letter - 1).value
                    if adjacent_cell == cell:  # Если совпадает с cell, добавляем в уникальные имена
                        unique_names.add(cell_in_row.value)

        return unique_names
# Обновленная функция для загрузки кнопок из Excel
def load_buttons_from_excel(file_path, sheet_name, column_letter):
    return list(get_unique_names_from_excel(file_path, sheet_name, column_letter))

# Функция для создания клавиатуры
def create_keyboard(buttons):
    keyboard = types.ReplyKeyboardMarkup(keyboard=[],resize_keyboard=True)
    for i in buttons:
        keyboard.keyboard.append([str(i)])
    return keyboard

crutch = 0
filenamesave = ""
allrow = ""
numberline=0
@dp.message(lambda message: message.text == "Изменить Отчёт")
async def exel_edit(message: types.Message, state: FSMContext):
    await message.answer("Занимаюсь работой с яндекс диском, подождите")
    update_files_from_yandex_disk("Отчёты/", "/tg test/Отчёты")
    file = open("admins.txt", 'r')
    booli = 0
    keyboard = types.ReplyKeyboardMarkup(keyboard=[], resize_keyboard=True)
    keyboard.keyboard.append([types.KeyboardButton(text="/back")])

    for i in file:
        if i == str(message.chat.id) + "\n":
            booli=1
    if booli==0:
        await message.answer("У вас недостаточно доступа")
        await state.clear()
    else:
        files = os.listdir("Отчёты")
        global lengh
        files.remove("Образец.xlsx")
        a = len(files)
        lengh = a
        text = ""
        await message.answer("Выберите отчёт\n", reply_markup=keyboard)
        for i in range(a):
            text = text+str(i + 1) + ". " + str(files[i]) + "\n"
        if text == "":
            await message.answer("Для начала нужно создать файл", reply_markup=greet_kb)
            await state.clear()
        else:
            await message.answer(text,reply_markup=keyboard)
            await state.set_state(editStates.waiting_for_chose_document)

@dp.message(editStates.waiting_for_chose_document)
async def chose_document_get(message: types.Message, state: FSMContext):
    if message.text == "/back":
        await message.answer("Вы возвращаетесь на главное меню",reply_markup=greet_kb)
        await state.clear()
    else:
        folder_path = 'Отчёты'
        files = os.listdir("Отчёты")
        files.remove("Образец.xlsx")
        global filenamesave
        file_name = message.text
        for i in range(lengh):
            if message.text in files[i]:
                file_name = message.text
            elif message.text == str(i+1):
                file_name = files[i]
        if os.path.exists(folder_path + "/" + file_name):
            filenamesave = file_name
            keyboard = types.ReplyKeyboardMarkup(keyboard=[], resize_keyboard=True)
            keyboard.keyboard.append([types.KeyboardButton(text="Изменить строку в отчёте")])
            keyboard.keyboard.append([types.KeyboardButton(text="Добавить строку")])
            keyboard.keyboard.append([types.KeyboardButton(text="/back")])
            await message.answer("Выберите режим работы: \n 1. Изменить строку в отчёте \n 2. Добавить строку", reply_markup=keyboard)
            await state.set_state(editStates.waiting_for_worktype)
        else:
            await message.answer("Ошибка файла с " + message.text + " номером/названием не существует, проверьте правильность номера/названия введёного вами")

@dp.message(editStates.waiting_for_worktype)
async def worktype_get(message: types.Message, state: FSMContext):
    if message.text in "Изменить строку в отчёте" or message.text == str(1):
        keyboard = types.ReplyKeyboardMarkup(keyboard=[], resize_keyboard=True)
        keyboard.keyboard.append([types.KeyboardButton(text="/back")])
        await message.answer("Напишите номер строки",reply_markup=keyboard)
        await state.set_state(editStates.waiting_for_chose_row)
    elif message.text in "Добавить строку" or message.text == str(2):
        global crutch
        crutch = 0
        keyboard = types.ReplyKeyboardMarkup(keyboard=[],resize_keyboard=True)
        keyboard.keyboard.append([types.KeyboardButton(text="Текущую дату")])
        keyboard.keyboard.append([types.KeyboardButton(text="Собственную дату")])
        keyboard.keyboard.append([types.KeyboardButton(text="/back")])
        await message.answer("Выберите вариант", reply_markup=keyboard)
        await state.set_state(editStates.waiting_for_chose_type_date)
    elif message.text == "/back":
        await message.answer("Вы возвращаетесь назак к выбору документа", reply_markup=ReplyKeyboardRemove())
        files = os.listdir("Отчёты")
        global lengh
        files.remove("Образец.xlsx")
        a = len(files)
        lengh = a
        text = ""
        await message.answer("Выберите отчёт\n", reply_markup=types.ReplyKeyboardRemove())
        for i in range(a):
            text = text + str(i + 1) + ". " + str(files[i]) + "\n"
        if text == "":
            await message.answer("Для начала нужно создать файл", reply_markup=greet_kb)
            await state.clear()
        else:
            await message.answer(text)
        await state.set_state(editStates.waiting_for_chose_document)
    else:
        await message.answer("Вы неправильно ввели команду")
        await state.set_state(editStates.waiting_for_worktype)


@dp.message(editStates.waiting_for_chose_row)
async def chose_row_get(message:types.Message, state:FSMContext):
    global crutch
    crutch = 0
    if message.text.isdigit():
        if not (int(message.text) <= 1):
            wb = openpyxl.load_workbook("Отчёты/" + filenamesave)
            sheet = wb.active
            if sheet.cell(row=int(message.text), column=1).value != None:
                global numberline
                numberline = int(message.text)
                row = get_allrow(numberline)
                wb = openpyxl.load_workbook("Отчёты/" + filenamesave)
                sheet = wb.active
                keyboard = types.ReplyKeyboardMarkup(keyboard=[], resize_keyboard=True)
                for col in range(1, 11):
                    # Берем значение из 1 строки и n строки для текущего столбца
                    cell_1 = sheet.cell(row=1, column=col).value
                    keyboard.keyboard.append([cell_1])
                keyboard.keyboard.append([types.KeyboardButton(text="/back")])
                await message.answer("Выберите вариант\n" + row, reply_markup=keyboard)
                await state.set_state(editStates.waiting_for_element_row)
            else:
                await message.answer("Простите но строка пустая, попробуйте ещё раз")
                await state.set_state(editStates.waiting_for_chose_row)
        else:
            await message.answer("Вы неправильно ввели номер строки, повторите попытку")
            await state.set_state(editStates.waiting_for_chose_row)
    elif (message.text == "/back"):
        await message.answer("Вы возвращаетесь назад к выбору режима работы")
        keyboard = types.ReplyKeyboardMarkup(keyboard=[], resize_keyboard=True)
        keyboard.keyboard.append([types.KeyboardButton(text="Изменить строку в отчёте")])
        keyboard.keyboard.append([types.KeyboardButton(text="Добавить строку")])
        keyboard.keyboard.append([types.KeyboardButton(text="/back")])
        await message.answer("Выберите режим работы: \n 1. Изменить строку в отчёте \n 2. Добавить строку",
                             reply_markup=keyboard)
        await state.set_state(editStates.waiting_for_worktype)
    else:
        await message.answer("Вы ввели не строку, попрогбуйте ещё раз")
        await state.set_state(editStates.waiting_for_chose_row)

@dp.message(editStates.waiting_for_element_row)
async def chose_element_get(message:types.Message, state:FSMContext):
    if message.text == "/back":
        await message.answer("Вы возвращаетесь к выбору номера строки")
        await message.answer("Напишите номер строки", reply_markup=ReplyKeyboardRemove())
        await state.set_state(editStates.waiting_for_chose_row)
    else:
        await message.answer("Удаление кнопок", reply_markup=ReplyKeyboardRemove())
        global crutch
        wb = openpyxl.load_workbook("Отчёты/" + filenamesave)
        sheet = wb.active
        alternating_values = set()
        for col in range(1, 11):
            # Берем значение из 1 строки и n строки для текущего столбца
            cell_1 = sheet.cell(row=1, column=col).value
            if cell_1 == message.text:
                crutch=col
        if crutch==2:
            await message.answer("Выберите категорию:")
            file = open("шаблон-файл.txt", "r", encoding="utf-8")
            exsemplenamesave = "None"
            for word in file:
                if filenamesave in word:
                    exsemplenamesave = word[:word.find("+-")]
            if exsemplenamesave != "None":
                buttons = load_buttons_from_excel("Шаблоны/" + exsemplenamesave, "Лист1", 1)
                keyboard = create_keyboard(buttons)
                keyboard.keyboard.append([KeyboardButton(text="/back")])
                await message.answer("Подгрузка кнопок", reply_markup=keyboard)
                await state.set_state(editStates.waiting_for_edit_row)
            else:
                await message.answer("Ошибка 404", reply_markup=greet_kb)
        elif crutch==3:
            await message.answer("Статью расходов:")
            file = open("шаблон-файл.txt", "r", encoding="utf-8")
            exsemplenamesave = "None"
            for word in file:
                if filenamesave in word:
                    exsemplenamesave = word[:word.find("+-")]
            if exsemplenamesave != "None":
                buttons = load_buttons_from_excel("Шаблоны/" + exsemplenamesave, "Лист1", 2)
                keyboard = create_keyboard(buttons)
                keyboard.keyboard.append([KeyboardButton(text="/back")])
                await message.answer("Подгрузка кнопок", reply_markup=keyboard)
                await state.set_state(editStates.waiting_for_edit_row)
        elif crutch==5:
            keyboard = types.ReplyKeyboardMarkup(keyboard=[], resize_keyboard=True)
            keyboard.keyboard.append([types.KeyboardButton(text="Грантовый")])
            keyboard.keyboard.append([types.KeyboardButton(text="Локальный")])
            keyboard.keyboard.append([types.KeyboardButton(text="/back")])
            await message.answer("Какой тип: \n 1. Грантовый \n 2. Локальный", reply_markup=keyboard)
            await state.set_state(editStates.waiting_for_edit_row)
        elif crutch == 8 or crutch == 9:
            await message.answer("Напишите колличество документов")
            await state.set_state(editStates.waiting_for_document_numbers)
        elif crutch==1:
            keyboard = types.ReplyKeyboardMarkup(keyboard=[], resize_keyboard=True)
            keyboard.keyboard.append([types.KeyboardButton(text="Текущую дату")])
            keyboard.keyboard.append([types.KeyboardButton(text="Собственную дату")])
            keyboard.keyboard.append([types.KeyboardButton(text="/back")])
            await message.answer("Выберите вариант", reply_markup=keyboard)
            await state.set_state(editStates.waiting_for_chose_type_date)
        else:
            await message.answer("Напишите текст")
            await state.set_state(editStates.waiting_for_edit_row)
@dp.message(editStates.waiting_for_edit_row)
async def edit_row_get(message:types.Message, state:FSMContext):
    if message.text == "/back":
        await message.answer("Вывозвращаетесь на шаг назад к выбору элемента", reply_markup=ReplyKeyboardRemove())
        row = get_allrow(numberline)
        wb = openpyxl.load_workbook("Отчёты/" + filenamesave)
        sheet = wb.active
        alternating_values = set()
        for col in range(1, 11):
            # Берем значение из 1 строки и n строки для текущего столбца
            cell_1 = sheet.cell(row=1, column=col).value
            alternating_values.add(cell_1)
        h = list(alternating_values)
        keyboard = create_keyboard(h)
        await message.answer("Выберите вариант\n" + row, reply_markup=keyboard)
        await state.set_state(editStates.waiting_for_element_row)
    else:
        wb = openpyxl.load_workbook("Отчёты/" + filenamesave)
        sheet = wb.active
        sheet.cell(row=numberline,column=crutch).value = message.text
        wb.save("Отчёты/" + filenamesave)
        row = get_allrow(numberline)
        keyboard = types.ReplyKeyboardMarkup(keyboard=[], resize_keyboard=True)
        keyboard.keyboard.append([types.KeyboardButton(text="Да")])
        keyboard.keyboard.append([types.KeyboardButton(text="Нет")])
        keyboard.keyboard.append([types.KeyboardButton(text="/back")])
        await message.answer("Строка изменена. \nНапишите Да для сохранения изменений\n" + row,reply_markup=ReplyKeyboardRemove())
        await message.answer("Подгрузка кнопок", reply_markup=keyboard)
        await state.set_state(editStates.waiting_for_accept_row)

@dp.message(editStates.waiting_for_chose_type_date)
async def type_gate(message:types.Message, state:FSMContext):
    if message.text == "Текущую дату":
        global filenamesave
        global numberline
        global allrow
        if crutch!=1:
            current_date = datetime.now().date()
            formatted_date = current_date.strftime("%d/%m/%Y")
            wb = openpyxl.load_workbook("Отчёты/" + filenamesave)
            sheet = wb.active
            i = 1
            A = sheet['A' + str(i)]
            while A.value != None:
                A = sheet['A' + str(i)]
                i += 1
            i = i - 1
            numberline = i
            A.value = formatted_date
            A.number_format = 'DD/MM/YYYY'
            allrow = "Дата:" + formatted_date + "\n"
            wb.save("Отчёты/" + filenamesave)
            await message.answer("Выберите категорию:", reply_markup=types.ReplyKeyboardRemove())
            file = open("шаблон-файл.txt", "r", encoding="utf-8")
            exsemplenamesave = "None"
            for word in file:
                if filenamesave in word:
                    exsemplenamesave = word[:word.find("+-")]
            if exsemplenamesave != "None":
                buttons = load_buttons_from_excel("Шаблоны/" + exsemplenamesave, "Лист1", 1)
                keyboard = create_keyboard(buttons)
                keyboard.keyboard.append([KeyboardButton(text="/back")])
                await message.answer("Подгрузка кнопок",reply_markup=keyboard)
            else:
                await message.answer("Ошибка 404", reply_markup=greet_kb)
            await state.set_state(editStates.waiting_for_chose_category)
        else:
            current_date = datetime.now().date()
            formatted_date = current_date.strftime("%d/%m/%Y")
            wb = openpyxl.load_workbook("Отчёты/" + filenamesave)
            sheet = wb.active
            A = sheet['A' + str(numberline)]
            A.value = formatted_date
            A.number_format = 'DD/MM/YYYY'
            allrow = "Дата:" + formatted_date + "\n"
            wb.save("Отчёты/" + filenamesave)
            row = get_allrow(numberline)
            keyboard = types.ReplyKeyboardMarkup(keyboard=[], resize_keyboard=True)
            keyboard.keyboard.append([types.KeyboardButton(text="Да")])
            keyboard.keyboard.append([types.KeyboardButton(text="Нет")])
            keyboard.keyboard.append([types.KeyboardButton(text="/back")])
            await message.answer("Строка изменена. \nНапишите Да для сохранения изменений\n" + row,
                                 reply_markup=ReplyKeyboardRemove())
            await message.answer("Подгрузка кнопок", reply_markup=keyboard)
            await state.set_state(editStates.waiting_for_accept_row)
    elif message.text == "Собственную дату":
        await message.answer("Введите дату (нужный формат: dd.mm.yyyy)")
        await state.set_state(editStates.waiting_for_chose_date)
    elif message.text=="/back":
        if crutch== 0:
            await message.answer("Вы возвращаетесь к выбору режима работы", reply_markup=ReplyKeyboardRemove())
            keyboard = types.ReplyKeyboardMarkup(keyboard=[], resize_keyboard=True)
            keyboard.keyboard.append([types.KeyboardButton(text="Изменить строку в отчёте")])
            keyboard.keyboard.append([types.KeyboardButton(text="Добавить строку")])
            keyboard.keyboard.append([types.KeyboardButton(text="/back")])
            await message.answer("Выберите режим работы: \n 1. Изменить строку в отчёте \n 2. Добавить строку",
                                 reply_markup=keyboard)
            await state.set_state(editStates.waiting_for_worktype)
        else:
            row = get_allrow(numberline)
            wb = openpyxl.load_workbook("Отчёты/" + filenamesave)
            sheet = wb.active
            keyboard = types.ReplyKeyboardMarkup(keyboard=[], resize_keyboard=True)
            for col in range(1, 11):
                # Берем значение из 1 строки и n строки для текущего столбца
                cell_1 = sheet.cell(row=1, column=col).value
                keyboard.keyboard.append([cell_1])
            keyboard.keyboard.append([types.KeyboardButton(text="/back")])
            await message.answer("Выберите вариант\n" + row, reply_markup=keyboard)
            await state.set_state(editStates.waiting_for_element_row)
@dp.message(editStates.waiting_for_chose_date)
async def date_gate(message:types.Message, state:FSMContext):
    if message.text == "Выход":
        await message.answer("Вы вышли")
        await state.clear()
    elif message.text=="/back":
        await message.answer("Вы возвращаетесь к выбору даты", reply_markup=ReplyKeyboardRemove())
        keyboard = types.ReplyKeyboardMarkup(keyboard=[], resize_keyboard=True)
        keyboard.keyboard.append([types.KeyboardButton(text="Текущую дату")])
        keyboard.keyboard.append([types.KeyboardButton(text="Собственную дату")])
        keyboard.keyboard.append([types.KeyboardButton(text="/back")])
        await message.answer("Выберите вариант", reply_markup=keyboard)
        await state.set_state(editStates.waiting_for_chose_type_date)
    else:
        date_object = convert_date_format(message.text)
        if date_object!="Неверный формат даты":
            global filenamesave
            global numberline
            global allrow
            if crutch!=1:
                wb = openpyxl.load_workbook("Отчёты/" + filenamesave)
                sheet = wb.active
                i = 1
                A = sheet['A' + str(i)]
                while A.value != None:
                    A = sheet['A' + str(i)]
                    i += 1
                i = i - 1
                numberline = i
                A.value = date_object
                A.number_format = 'DD/MM/YYYY'
                allrow = "Дата:" + str(date_object) + "\n"
                wb.save("Отчёты/" + filenamesave)
                await message.answer("Выберите категорию:")
                file = open("шаблон-файл.txt", "r", encoding="utf-8")
                exsemplenamesave = "None"
                for word in file:
                    if filenamesave in word:
                        exsemplenamesave = word[:word.find("+-")]
                if exsemplenamesave != "None":
                    buttons = load_buttons_from_excel("Шаблоны/" + exsemplenamesave, "Лист1", 1)
                    keyboard = create_keyboard(buttons)
                    keyboard.keyboard.append([KeyboardButton(text="/back")])
                    await message.answer("Подгрузка кнопок", reply_markup=keyboard)
                    await state.set_state(editStates.waiting_for_chose_category)
                else:
                    await message.answer("Ошибка 404", reply_markup=greet_kb)
            else:
                wb = openpyxl.load_workbook("Отчёты/" + filenamesave)
                sheet = wb.active
                A = sheet['A' + str(numberline)]
                A.value = date_object
                A.number_format = 'DD/MM/YYYY'
                allrow = "Дата:" + str(date_object) + "\n"
                wb.save("Отчёты/" + filenamesave)
                row = get_allrow(numberline)
                keyboard = types.ReplyKeyboardMarkup(keyboard=[], resize_keyboard=True)
                keyboard.keyboard.append([types.KeyboardButton(text="Да")])
                keyboard.keyboard.append([types.KeyboardButton(text="Нет")])
                keyboard.keyboard.append([types.KeyboardButton(text="/back")])
                await message.answer("Строка изменена. \nНапишите Да для сохранения изменений\n" + row,
                                     reply_markup=ReplyKeyboardRemove())
                await message.answer("Подгрузка кнопок", reply_markup=keyboard)
                await state.set_state(editStates.waiting_for_accept_row)
        else:
            await message.answer("Вы не правильно ввели дату, повторите ещё раз (нужный формат: dd.mm.yyyy)")
            await state.set_state(editStates.waiting_for_chose_date)

@dp.message(editStates.waiting_for_chose_category)
async def category_get(message:types.Message, state: FSMContext):
    if message.text=="/back":
        await message.answer("Вы вернулись на один шаг назад",reply_markup=ReplyKeyboardRemove())
        keyboard = types.ReplyKeyboardMarkup(keyboard=[], resize_keyboard=True)
        keyboard.keyboard.append([types.KeyboardButton(text="Текущую дату")])
        keyboard.keyboard.append([types.KeyboardButton(text="Собственную дату")])
        keyboard.keyboard.append([types.KeyboardButton(text="/back")])
        await message.answer("Выберите вариант", reply_markup=keyboard)
        await state.set_state(editStates.waiting_for_chose_type_date)
    elif message.text == "Выход":
        wb = openpyxl.load_workbook("Отчёты/" + filenamesave)
        sheet = wb.active
        sheet.delete_rows(numberline)
        wb.save("Отчёты/" + filenamesave)
        await message.answer("Всё удалено, кроме документов на яндекс диске, пожалуйста не забудьте их удалить")
        await state.clear()
    else:

        wb = openpyxl.load_workbook("Отчёты/" + filenamesave)
        sheet = wb.active
        B = sheet['B' + str(numberline)]
        B.value = message.text
        global allrow
        allrow = allrow + "Категория:" + message.text + "\n"
        wb.save("Отчёты/" + filenamesave)
        await message.answer("Статью расходов:")
        file = open("шаблон-файл.txt", "r", encoding="utf-8")
        exsemplenamesave = "None"
        for word in file:
            if filenamesave in word:
                exsemplenamesave = word[:word.find("+-")]
        if exsemplenamesave != "None":
            buttons = load_buttons_from_excel("Шаблоны/" + exsemplenamesave, "Лист1", 2)
            keyboard = create_keyboard(buttons)
            keyboard.keyboard.append([KeyboardButton(text="/back")])
            await message.answer("Подгрузка кнопок", reply_markup=keyboard)
            await state.set_state(editStates.waiting_for_chose_expense_item)
        else:
            await message.answer("Ошибка 404", reply_markup=greet_kb)

@dp.message(editStates.waiting_for_chose_expense_item)
async def expencse_get(message:types.Message, state: FSMContext):
    if message.text=="/back":
        file = open("шаблон-файл.txt", "r", encoding="utf-8")
        exsemplenamesave = "None"
        for word in file:
            if filenamesave in word:
                exsemplenamesave = word[:word.find("+-")]
        if exsemplenamesave != "None":
            await message.answer("Вы вернулись на один шаг назад к выбору категории")
            buttons = load_buttons_from_excel("Шаблоны/" + exsemplenamesave, "Лист1", 1)
            keyboard = create_keyboard(buttons)
            keyboard.keyboard.append([KeyboardButton(text="/back")])
            await message.answer("Подгрузка кнопок", reply_markup=keyboard)
            await state.set_state(editStates.waiting_for_chose_category)
    elif message.text == "Выход":
        wb = openpyxl.load_workbook("Отчёты/" + filenamesave)
        sheet = wb.active
        sheet.delete_rows(numberline)
        wb.save("Отчёты/" + filenamesave)
        await message.answer("Всё удалено, кроме документов на яндекс диске, пожалуйста не забудьте их удалить")
        await state.clear()
    else:
        wb = openpyxl.load_workbook("Отчёты/" + filenamesave)
        sheet = wb.active
        C = sheet['C' + str(numberline)]
        global allrow
        allrow = allrow + "Статья расходов:" + message.text + "\n"
        C.value = message.text
        wb.save("Отчёты/" + filenamesave)
        if crutch!=2 and crutch!=3:
            await message.answer("Напишите название",reply_markup=types.ReplyKeyboardRemove())
            await state.set_state(editStates.waiting_for_chose_name)
        else:
            allrow = get_allrow(numberline)
            keyboard = types.ReplyKeyboardMarkup(keyboard=[], resize_keyboard=True)
            keyboard.keyboard.append([types.KeyboardButton(text="Да")])
            keyboard.keyboard.append([types.KeyboardButton(text="Нет")])
            keyboard.keyboard.append([types.KeyboardButton(text="/back")])
            await message.answer(allrow)
            await message.answer(
                "Всё верно? \nНапишите Да, если всё ок, либо вернитесь и справьте нужную строку, либо напишите Выход, чтобы удалить последние изменения в файле",
                reply_markup=keyboard)
            await state.set_state(editStates.waiting_for_accept_row)

@dp.message(editStates.waiting_for_chose_name)
async def name_get(message:types.Message, state:FSMContext):
    if message.text=="/back":
        await message.answer("Вы вернулись на один шаг назад к выбору статьи расходов",reply_markup=ReplyKeyboardRemove())
        file = open("шаблон-файл.txt", "r", encoding="utf-8")
        exsemplenamesave = "None"
        for word in file:
            if filenamesave in word:
                exsemplenamesave = word[:word.find("+-")]
        if exsemplenamesave != "None":
            buttons = load_buttons_from_excel("Шаблоны/" + exsemplenamesave, "Лист1", 2)
            keyboard = create_keyboard(buttons)
            keyboard.keyboard.append([KeyboardButton(text="/back")])
            await message.answer("Подгрузка кнопок", reply_markup=keyboard)
            await state.set_state(editStates.waiting_for_chose_expense_item)
        else:
            await message.answer("Ошибка 404", reply_markup=greet_kb)
        await state.set_state(editStates.waiting_for_chose_expense_item)
    elif message.text == "Выход":
        wb = openpyxl.load_workbook("Отчёты/" + filenamesave)
        sheet = wb.active
        sheet.delete_rows(numberline)
        wb.save("Отчёты/" + filenamesave)
        await message.answer("Всё удалено, кроме документов на яндекс диске, пожалуйста не забудьте их удалить")
        await state.clear()
    else:
        wb = openpyxl.load_workbook("Отчёты/" + filenamesave)
        sheet = wb.active
        D = sheet['D' + str(numberline)]
        global allrow
        allrow = allrow + "Тип:" + message.text + "\n"
        D.value = message.text
        wb.save("Отчёты/" + filenamesave)
        keyboard = types.ReplyKeyboardMarkup(keyboard=[], resize_keyboard=True)
        keyboard.keyboard.append([types.KeyboardButton(text="Грантовый")])
        keyboard.keyboard.append([types.KeyboardButton(text="Локальный")])
        keyboard.keyboard.append([types.KeyboardButton(text="/back")])
        await message.answer("Какой тип: \n 1. Грантовый \n 2. Локальный", reply_markup=keyboard)
        await state.set_state(editStates.waiting_for_type)

@dp.message(editStates.waiting_for_type)
async def type_get(message:types.Message,state:FSMContext):
    global allrow
    if message.text=="/back":
        await message.answer("Вы вернулись на один шаг назад к выбору названия",reply_markup=ReplyKeyboardRemove())
        await state.set_state(editStates.waiting_for_chose_name)
    elif message.text == "Выход":
        wb = openpyxl.load_workbook("Отчёты/" + filenamesave)
        sheet = wb.active
        sheet.delete_rows(numberline)
        wb.save("Отчёты/" + filenamesave)
        await message.answer("Всё удалено, кроме документов на яндекс диске, пожалуйста не забудьте их удалить")
        await state.clear()
    else:
        if message.text == "1" or message.text == "Грантовый":
            wb = openpyxl.load_workbook("Отчёты/" + filenamesave)
            sheet = wb.active
            E = sheet['E' + str(numberline)]
            allrow = allrow + "Тип:" + "Грантовый" + "\n"
            E.value = message.text
            wb.save("Отчёты/" + filenamesave)
            await message.answer("Введите колличество",reply_markup=types.ReplyKeyboardRemove())
            await state.set_state(editStates.waiting_for_chose_count)
        elif message.text == "2" or message.text == "Локальный":
            wb = openpyxl.load_workbook("Отчёты/" + filenamesave)
            sheet = wb.active
            E = sheet['E' + str(numberline)]
            allrow = allrow + "Тип:" + "Локальный" + "\n"
            E.value = message.text
            wb.save("Отчёты/" + filenamesave)
            await message.answer("Введите колличество",reply_markup=types.ReplyKeyboardRemove())
            await state.set_state(editStates.waiting_for_chose_count)
        else:
            await message.answer("Вы неправильно ввели тип, попробуйте ещё раз")
            await state.set_state(editStates.waiting_for_type)

@dp.message(editStates.waiting_for_chose_count)
async def count_get(message:types.Message, state: FSMContext):
    if message.text=="/back":
        await message.answer("Вы вернулись на один шаг назад к выбору типа")
        keyboard = types.ReplyKeyboardMarkup(keyboard=[], resize_keyboard=True)
        keyboard.keyboard.append([types.KeyboardButton(text="Грантовый")])
        keyboard.keyboard.append([types.KeyboardButton(text="Локальный")])
        keyboard.keyboard.append([types.KeyboardButton(text="/back")])
        await message.answer("Какой тип: \n 1. Грантовый \n 2. Локальный", reply_markup=keyboard)
        await state.set_state(editStates.waiting_for_type)
    elif message.text == "Выход":
        wb = openpyxl.load_workbook("Отчёты/" + filenamesave)
        sheet = wb.active
        sheet.delete_rows(numberline)
        wb.save("Отчёты/" + filenamesave)
        await message.answer("Всё удалено, кроме документов на яндекс диске, пожалуйста не забудьте их удалить", reply_markup=greet_kb)
        await state.clear()
    else:
        wb = openpyxl.load_workbook("Отчёты/" + filenamesave)
        sheet = wb.active
        F = sheet['F' + str(numberline)]
        global allrow
        allrow = allrow + "Количество:" + message.text + "\n"
        F.value = message.text
        wb.save("Отчёты/" + filenamesave)
        await message.answer("Напишите сумму")
        await state.set_state(editStates.waiting_for_chose_price)

@dp.message(editStates.waiting_for_chose_price)
async def price_get(message:types.Message, state:FSMContext):
    if message.text=="/back":
        await message.answer("Вы вернулись на один шаг назад к выбору количества ")
        await state.set_state(editStates.waiting_for_chose_count)
    elif message.text == "Выход":
        wb = openpyxl.load_workbook("Отчёты/" + filenamesave)
        sheet = wb.active
        sheet.delete_rows(numberline)
        wb.save("Отчёты/" + filenamesave)
        await message.answer("Всё удалено, кроме документов на яндекс диске, пожалуйста не забудьте их удалить", reply_markup=greet_kb)
        await state.clear()
    else:
        wb = openpyxl.load_workbook("Отчёты/" + filenamesave)
        sheet = wb.active
        G = sheet['G' + str(numberline)]
        global allrow
        allrow = allrow + "Сумма:" + message.text + "\n"
        G.value = message.text
        wb.save("Отчёты/" + filenamesave)
        await message.answer("Напишите колличество документов")
        await state.set_state(editStates.waiting_for_document_numbers)

document_number = 0
document_count = 0
@dp.message(editStates.waiting_for_document_numbers)
async def number_get(message:types.Message, state:FSMContext):
    if message.text=="/back":
        if crutch==0:
            await message.answer("Вы вернулись на один шаг назад к выбору цены")
            await state.set_state(editStates.waiting_for_chose_price)
        else:
            row = get_allrow(numberline)
            wb = openpyxl.load_workbook("Отчёты/" + filenamesave)
            sheet = wb.active
            keyboard = types.ReplyKeyboardMarkup(keyboard=[], resize_keyboard=True)
            for col in range(1, 11):
                # Берем значение из 1 строки и n строки для текущего столбца
                cell_1 = sheet.cell(row=1, column=col).value
                keyboard.keyboard.append([cell_1])
            keyboard.keyboard.append([types.KeyboardButton(text="/back")])
            await message.answer("Выберите вариант\n" + row, reply_markup=keyboard)
            await state.set_state(editStates.waiting_for_element_row)
    elif message.text == "Выход" and crutch==0:
        wb = openpyxl.load_workbook("Отчёты/" + filenamesave)
        sheet = wb.active
        sheet.delete_rows(numberline)
        wb.save("Отчёты/" + filenamesave)
        await message.answer("Всё удалено, кроме документов на яндекс диске, пожалуйста не забудьте их удалить", reply_markup=greet_kb)
        await state.clear()
    else:
        global document_number, document_count
        document_number=int(message.text)
        document_count=0
        wb = openpyxl.load_workbook("Отчёты/" + filenamesave)
        sheet = wb.active
        I = sheet['I' + str(numberline)]
        I.value=""
        wb.save("Отчёты/" + filenamesave)
        if document_number > 0:
            file = open("documenttypes.txt", "r", encoding="utf-8")
            doctypes = ""
            for i in file:
                doctypes = doctypes+i
            await message.answer("Напишите названия документов\n" + doctypes+"\n"+"7. Свой вариант")
            await state.set_state(editStates.waiting_for_document_types)
        else:
            await message.answer("Есть какие-то комментарии?\n Если нет поставьте -")
            await state.set_state(editStates.waiting_for_comments)
checko = 0
doctext = ""
@dp.message(editStates.waiting_for_document_types)
async def types_get(message:types.Message, state:FSMContext):
    if message.text=="/back":
        await message.answer("Вы вернулись на один шаг назад к выбору количества документов")
        await state.set_state(editStates.waiting_for_document_numbers)
    elif message.text == "Выход" and crutch==0:
        wb = openpyxl.load_workbook("Отчёты/" + filenamesave)
        sheet = wb.active
        sheet.delete_rows(numberline)
        wb.save("Отчёты/" + filenamesave)
        await message.answer("Всё удалено, кроме документов на яндекс диске, пожалуйста не забудьте их удалить", reply_markup=greet_kb)
        await state.clear()
    else:
        file = open("documenttypes.txt", "r", encoding="utf-8")
        text = message.text
        global checko
        global doctext
        if len(text) == document_number or checko%7==0:
            if checko%7!=0 or checko==0:
                for i in file:
                    for g in range(document_number):
                        if text[g] in i:
                            doctext=doctext+i[i.find(" ")+1:] + " "
                for g in range(document_number):
                    if text[g] == "7":
                        checko=checko+7
            else:
                doctext=doctext+" "+message.text
                checko=checko-7
                if checko==0:
                    checko=1
            if checko%7!=0 or checko==0:
                wb = openpyxl.load_workbook("Отчёты/" + filenamesave)
                sheet = wb.active
                H = sheet['H' + str(numberline)]
                global allrow
                allrow = allrow + "Документы:" + doctext + "\n"
                H.value = doctext
                wb.save("Отчёты/" + filenamesave)
                checko=0
                keyboard = types.ReplyKeyboardMarkup(keyboard=[], resize_keyboard=True)
                keyboard.keyboard.append([types.KeyboardButton(text="/back")])
                await message.answer("Можете начать отправлять документы, только по одному, а то я могу не успеть:(",
                                     reply_markup=keyboard)
                await state.set_state(editStates.waiting_for_document)
                doctext=""
            else:
                await message.answer("Напишите свой вариант")
                await state.set_state(editStates.waiting_for_document_types)
        else:
            await message.answer("Вы ошиблись либо с количеством документов, либо с количеством типов документов, попробуйте ещё раз")
            await state.set_state(editStates.waiting_for_document_types)

@dp.message(editStates.waiting_for_document)
async def document_get(message:types.Message, state:FSMContext):
    global document_count
    if message.text=="/back":
        await message.answer("Вы вернулись на один шаг назад к выбору типа документа")
        await state.set_state(editStates.waiting_for_document_types)
    elif message.text == "Выход" and crutch==0:
        wb = openpyxl.load_workbook("Отчёты/" + filenamesave)
        sheet = wb.active
        sheet.delete_rows(numberline)
        wb.save("Отчёты/" + filenamesave)
        await message.answer("Всё удалено, кроме документов на яндекс диске, пожалуйста не забудьте их удалить", reply_markup=greet_kb)
        await state.clear()
    elif document_number==document_count:
        await message.answer("Все файлы загружены, есть какие-то комментарии?\n Если нет поставьте -")
        await state.set_state(editStates.waiting_for_comments)
    else:
        if message.document is not None and message.document.file_name is not None:
            document_count=document_count+1
            document = message.document
            file_id = document.file_id
            file_info = await bot.get_file(file_id)
            file_path = file_info.file_path
            file_name = document.file_name
            file_path_to_save = os.path.join("Документы", file_name)
            downloaded_file = await bot.download_file(file_path)
            with open(file_path_to_save, "wb") as f:
                f.write(downloaded_file.read())
            await message.answer("Занимаюсь работой с яндекс диском, подождите")
            y.upload("Документы/" + file_name, "/tg test/Документы/" +file_name, overwrite=True)
            public_url = y.get_download_link("/tg test/Документы/" + file_name)
            wb = openpyxl.load_workbook("Отчёты/" + filenamesave)
            sheet = wb.active
            I = sheet['I' + str(numberline)]
            global allrow
            allrow = allrow + "Документы:" + public_url + "\n"
            if I.value != None:
                I.value = str(I.value) + public_url + "\n"
            else:
                I.value = public_url + "\n"
            wb.save("Отчёты/" + filenamesave)
            if document_number== document_count:
                await message.answer("Все файлы загружены, есть какие-то комментарии?\nЕсли нет поставьте -")
                await state.set_state(editStates.waiting_for_comments)
            else:
                await message.answer("Принял, ожидайте")
                await message.answer("Присылайте следующий документ")
                await state.set_state(editStates.waiting_for_document)
        else:
            await message.answer("Нужно присылать документы так же, изображения должны быть отправленны, как документ,а если вы ошиблись в колличестве документов, то нажмите Назад")
            await state.set_state(editStates.waiting_for_document)
@dp.message(editStates.waiting_for_comments)
async def comments_get(message:types.Message, state: FSMContext):
    global allrow
    if message.text == "-":
        if crutch!=0:
            allrow=get_allrow(numberline)
            keyboard = types.ReplyKeyboardMarkup(keyboard=[], resize_keyboard=True)
            keyboard.keyboard.append([types.KeyboardButton(text="Да")])
            keyboard.keyboard.append([types.KeyboardButton(text="Нет")])
            keyboard.keyboard.append([types.KeyboardButton(text="/back")])
            await message.answer(allrow)
            await message.answer(
                "Всё верно? \nНапишите Да, если всё ок, либо вернитесь и справьте нужную строку, либо напишите Выход, чтобы удалить последние изменения в файле",
                reply_markup=keyboard)
            await state.set_state(editStates.waiting_for_accept_row)
        else:
            keyboard = types.ReplyKeyboardMarkup(keyboard=[], resize_keyboard=True)
            keyboard.keyboard.append([types.KeyboardButton(text="Да")])
            keyboard.keyboard.append([types.KeyboardButton(text="Нет")])
            keyboard.keyboard.append([types.KeyboardButton(text="/back")])
            await message.answer(allrow)
            await message.answer("Всё верно? \nНапишите Да, если всё ок, либо вернитесь и справьте нужную строку, либо напишите Выход, чтобы удалить последние изменения в файле",reply_markup=keyboard)
            await state.set_state(editStates.waiting_for_accept_row)
    elif message.text == "Выход":
        wb = openpyxl.load_workbook("Отчёты/" + filenamesave)
        sheet = wb.active
        sheet.delete_rows(numberline)
        wb.save("Отчёты/" + filenamesave)
        await message.answer("Всё удалено, кроме документов на яндекс диске, пожалуйста не забудьте их удалить", reply_markup=greet_kb)
        await state.clear()
    elif message.text=="/back":
        await message.answer("Вы вернулись на один шаг назад, к выбору количества документов")
        await state.set_state(editStates.waiting_for_document_numbers)
    else:
        keyboard = types.ReplyKeyboardMarkup(keyboard=[], resize_keyboard=True)
        keyboard.keyboard.append([types.KeyboardButton(text="Да")])
        keyboard.keyboard.append([types.KeyboardButton(text="Нет")])
        keyboard.keyboard.append([types.KeyboardButton(text="/back")])
        wb = openpyxl.load_workbook("Отчёты/" + filenamesave)
        sheet = wb.active
        I = sheet['I' + str(numberline)]
        allrow = allrow + "Комментарии:" + message.text
        I.value = str(I.value) + message.text
        wb.save("Отчёты/" + filenamesave)
        if crutch!=0:
            allrow=get_allrow(numberline)
        await message.answer(allrow)
        await message.answer("Всё верно? \nНапишите Да, если всё ок, либо вернитесь и справьте нужную строку, либо напишите Выход, чтобы удалить последние изменения в файле",reply_markup=keyboard)
        await state.set_state(editStates.waiting_for_accept_row)
@dp.message(editStates.waiting_for_accept_row)
async def accept_get(message:types.Message, state: FSMContext):
    if message.text == "Да":
        if crutch== 0:
            await message.answer("Занимаюсь работой с яндекс диском, подождите")
            y.upload("Отчёты/" + filenamesave, "/tg test/Отчёты/" + filenamesave, overwrite=True)
            await message.answer("Супер, файлы загружены на диск", reply_markup=greet_kb)
            await state.clear()
        else:
            await message.answer("Занимаюсь работой с яндекс диском, подождите")
            y.upload("Отчёты/" + filenamesave, "/tg test/Отчёты/" + filenamesave, overwrite=True)
            keyboard = types.ReplyKeyboardMarkup(keyboard=[], resize_keyboard=True)
            keyboard.keyboard.append([types.KeyboardButton(text="Да")])
            keyboard.keyboard.append([types.KeyboardButton(text="Нет")])
            keyboard.keyboard.append([types.KeyboardButton(text="Я хочу продолжить редактирование строки")])
            await message.answer("Супер, файлы загружены на диск")
            await message.answer("Вы хотите продолжить редактирование этого документа или нет?", reply_markup=keyboard)
            await state.set_state(editStates.waiting_for_chose)
    elif message.text == "Выход" and crutch==0:
        wb = openpyxl.load_workbook("Отчёты/" + filenamesave)
        sheet = wb.active
        sheet.delete_rows(numberline)
        wb.save("Отчёты/" + filenamesave)
        await message.answer("Всё удалено, кроме документов на яндекс диске, пожалуйста не забудьте их удалить", reply_markup=greet_kb)
        await state.clear()
    elif message.text=="/back":
        await message.answer("Вы вернулись на один шаг назад, к добавлению комментариев", reply_markup=ReplyKeyboardRemove())
        await state.set_state(editStates.waiting_for_comments)
    elif message.text=="Нет":
        if crutch!=0:
            remote_path = f"{"/tg test/Отчёты"}/{filenamesave}"
            local_path = os.path.join("Отчёты/", filenamesave)
            y.download(remote_path, local_path)
        row = get_allrow(numberline)
        wb = openpyxl.load_workbook("Отчёты/" + filenamesave)
        sheet = wb.active
        keyboard = types.ReplyKeyboardMarkup(keyboard=[], resize_keyboard=True)
        for col in range(1, 11):
            # Берем значение из 1 строки и n строки для текущего столбца
            cell_1 = sheet.cell(row=1, column=col).value
            keyboard.keyboard.append([cell_1])
        keyboard.keyboard.append([types.KeyboardButton(text="/back")])
        await message.answer("Удаление кнопок",reply_markup=keyboard)
        await message.answer("Выберите вариант\n" + row, reply_markup=keyboard)
        await state.set_state(editStates.waiting_for_element_row)
@dp.message(editStates.waiting_for_chose)
async def chossse_get(message:types.Message,state:FSMContext):
    if message.text=="Да":
        await message.answer("Возвращаю к выбору режима работы", reply_markup=ReplyKeyboardRemove())
        keyboard = types.ReplyKeyboardMarkup(keyboard=[], resize_keyboard=True)
        keyboard.keyboard.append([types.KeyboardButton(text="Изменить строку в отчёте")])
        keyboard.keyboard.append([types.KeyboardButton(text="Добавить строку")])
        keyboard.keyboard.append([types.KeyboardButton(text="/back")])
        await message.answer("Выберите режим работы: \n 1. Изменить строку в отчёте \n 2. Добавить строку",
                             reply_markup=keyboard)
        await state.set_state(editStates.waiting_for_worktype)
    elif str(message.text)=="Я хочу продолжить редактирование строки":
        row = get_allrow(numberline)
        wb = openpyxl.load_workbook("Отчёты/" + filenamesave)
        sheet = wb.active
        keyboard = types.ReplyKeyboardMarkup(keyboard=[], resize_keyboard=True)
        for col in range(1, 11):
            # Берем значение из 1 строки и n строки для текущего столбца
            cell_1 = sheet.cell(row=1, column=col).value
            keyboard.keyboard.append([cell_1])
        keyboard.keyboard.append([types.KeyboardButton(text="/back")])
        await message.answer("Выберите вариант\n" + row, reply_markup=keyboard)
        await state.set_state(editStates.waiting_for_element_row)
    else:
        await message.answer("Возвращаю вас к началу", reply_markup=greet_kb)
        await state.clear()